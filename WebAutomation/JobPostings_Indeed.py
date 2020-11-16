# Create a program to analyze desired qualifications
# for entry level tech positions on Indeed

from selenium import webdriver
from selenium.webdriver.common.by import By 
from selenium.webdriver.support import expected_conditions as EC 
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait

from bs4 import BeautifulSoup
import time
import pandas as pd
import xlsxwriter
from openpyxl import Workbook
from openpyxl import load_workbook

options = webdriver.ChromeOptions()
options.add_argument('--ignore-certificate-errors')
options.add_argument('--incognito')
driver = webdriver.Chrome(executable_path="/mnt/c/Users/Public/chromedriver_win32 (1)/chromedriver.exe", chrome_options=options)
driver.get("https://www.indeed.com/jobs?q=software+engineer+entry+level&l=")

# Wait for page to load and find the job posting card
timeout = 30
try:
    WebDriverWait(driver, timeout).until(EC.visibility_of_element_located((By.ID, "resultsCol")))
except TimeoutException:
    driver.quit()

# Retrieve job card preview information
"""
list_table_elements = driver.find_element(By.XPATH, "//table[@id='pageContent']/tbody/tr//td[@id='resultsCol']")
job_card = list_table_elements.find_elements(By.CLASS_NAME, "jobsearch-SerpJobCard")
for card in job_card:
    print(card.text + "\n")
"""

# Full Job Description
"""
elements = driver.find_elements_by_class_name('jobsearch-SerpJobCard')
# job_container = driver.find_elements(By.ID, 'vjs-container-iframe')
page_source = driver.page_source
soup = BeautifulSoup(page_source, 'html.parser')
# job = soup.find_all('div', class_='jobsearch-SerpJobCard')
# print(len(job))
for link in elements:
    webdriver.ActionChains(driver).move_to_element(link).click(link).perform()
    job_container = soup.find(id='vjs-container')
    job_description = job_container.find('ul').find('li').get_text()
    print(job_description)
    time.sleep(1)
"""

# Append basic job details
job_positions = []
companies = []
# companies = set()
location = []
job_overview = []
elements = driver.find_elements_by_class_name('jobsearch-SerpJobCard')
for job_preview in elements:
    # webdriver.ActionChains(driver).move_to_element(job_preview).click(job_preview).perform()
    job_positions.append(job_preview.find_element_by_class_name('title').text)
    companies.append(job_preview.find_element_by_class_name('company').text)
    location.append(job_preview.find_element_by_class_name('location').text)
    job_overview.append(job_preview.find_element_by_class_name('summary').text)
    time.sleep(1)

# Store data in csv file
data = {'Location': location, 'Company': companies, 'Position': job_positions, 'Job Overview': job_overview}
# print(len(location), len(companies), len(job_positions), len(job_overview))
"""
df = pd.DataFrame.from_dict(data)
df.to_csv('job_description.csv', encoding='utf-8')
"""
"""
df = pd.DataFrame.from_dict(data)
writer = pd.ExcelWriter('output.xlsx')
df.to_excel(writer)
writer.save()
"""
sheet_name = set(companies)
wb = Workbook()
dest_filename = 'output.xlsx'

for sheet in sheet_name:
    wb.create_sheet(title=sheet)
wb.save(filename=dest_filename)
print('DataFrame was written successfully to Excel File')
# writer = pd.ExcelWriter('job_postings.xlsx', engine='xlsxwriter', encoding='utf-8')
# for sheet_name in companies:
#     sheet_name.to_excel(writer, sheet_name=sheet_name, index=False)
# writer.save()
